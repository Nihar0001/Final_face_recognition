from flask import Flask, render_template, request, redirect, url_for, session, flash
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
import face_recognition
import cv2
import numpy as np
import os
import openpyxl
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
import base64
import re

# Flask app configuration
app = Flask(__name__)
app.secret_key = "your_secret_key"
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///attendance_system.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = 'known_faces'  # Directory to store facial images
app.config['ALLOWED_EXTENSIONS'] = {'png', 'jpg', 'jpeg'}
app.config['STUDENT_FOLDER'] = 'student_faces'
app.config['ATTENDANCE_FOLDER'] = 'attendance_records'

db = SQLAlchemy(app)

# Ensure directories exist
for folder in [app.config['UPLOAD_FOLDER'], app.config['STUDENT_FOLDER'], app.config['ATTENDANCE_FOLDER']]:
    if not os.path.exists(folder):
        os.makedirs(folder)

# User Model
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(120), nullable=False)
    role = db.Column(db.String(20), nullable=False)  # 'student' or 'staff'
    subject = db.Column(db.String(80), nullable=True)  # Only for staff

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

class Student(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    full_name = db.Column(db.String(120), nullable=False)
    roll_number = db.Column(db.String(20), nullable=False)
    class_name = db.Column(db.String(20), nullable=False)

# Function to check allowed file extensions
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

# Load known faces
known_face_encodings = []
known_face_names = []

def load_known_faces():
    global known_face_encodings, known_face_names
    known_face_encodings = []
    known_face_names = []

    for image_name in os.listdir(app.config['STUDENT_FOLDER']):
        image_path = os.path.join(app.config['STUDENT_FOLDER'], image_name)
        image = face_recognition.load_image_file(image_path)
        face_encoding = face_recognition.face_encodings(image)

        if len(face_encoding) > 0:
            known_face_encodings.append(face_encoding[0])
            known_face_names.append(os.path.splitext(image_name)[0])
        else:
            print(f"No face found in {image_name}")

load_known_faces()  # Load faces at startup

@app.route("/")
def home():
    return render_template("index.html")

@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]
        role = request.form["role"]

        if User.query.filter_by(username=username).first():
            flash("Username already exists. Please choose a different one.")
            return redirect(url_for("register"))

        new_user = User(username=username, role=role)

        if role == "staff":
            subject = request.form.get("subject")
            if not subject:
                flash("Please select a subject for staff members.")
                return redirect(url_for("register"))
            new_user.subject = subject

        new_user.set_password(password)
        db.session.add(new_user)

        if role == "student":
            # Create student profile
            full_name = request.form["full_name"]
            roll_number = request.form["roll_number"]
            class_name = request.form["class_name"]

            new_student = Student(
                username=username,
                full_name=full_name,
                roll_number=roll_number,
                class_name=class_name
            )
            db.session.add(new_student)

        db.session.commit()
        flash("Registration successful. Please login.")
        return redirect(url_for("login"))

    return render_template("register.html")

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]

        user = User.query.filter_by(username=username).first()

        if user and user.check_password(password):
            session["username"] = username
            session["role"] = user.role
            if user.role == "staff":
                session["subject"] = user.subject
            flash("Login successful!")
            return redirect(url_for("dashboard"))
        else:
            flash("Invalid credentials. Please try again.")
            return redirect(url_for("login"))

    return render_template("login.html")

@app.route("/dashboard")
def dashboard():
    if "username" not in session:
        return redirect(url_for("login"))

    if session["role"] == "student":
        # Check if student has uploaded face
        face_path = os.path.join(app.config['STUDENT_FOLDER'], f"{session['username']}.jpg")
        if not os.path.exists(face_path):
            flash("Please upload your facial image.")
            return redirect(url_for("upload_face"))

    return render_template("dashboard.html", username=session["username"], role=session["role"])

@app.route("/upload_face", methods=["GET", "POST"])
def upload_face():
    if "username" not in session or session["role"] != "student":
        return redirect(url_for("login"))

    if request.method == "POST" and 'file' in request.files:
        file = request.files["file"]
        if file.filename == "":
            flash("No selected file")
            return redirect(request.url)
        if file and allowed_file(file.filename):
            filename = secure_filename(f"{session['username']}.jpg")
            filepath = os.path.join(app.config["STUDENT_FOLDER"], filename)
            file.save(filepath)
            load_known_faces()
            flash("Face data uploaded successfully!")
            return redirect(url_for("dashboard"))

    return render_template("upload_face.html")

@app.route("/upload_face_camera", methods=["POST"])
def upload_face_camera():
    if "username" not in session or session["role"] != "student":
        return redirect(url_for("login"))

    camera_image_data = request.form.get('camera_image_data')
    if not camera_image_data:
        flash("No image data received from camera.")
        return redirect(url_for("upload_face"))

    try:
        # Extract base64 data
        image_data = re.sub('^data:image/.+;base64,', '', camera_image_data)
        image_bytes = base64.b64decode(image_data)

        filename = secure_filename(f"{session['username']}.jpg")
        filepath = os.path.join(app.config["STUDENT_FOLDER"], filename)

        with open(filepath, 'wb') as f:
            f.write(image_bytes)

        load_known_faces()
        flash("Face data captured from camera and saved successfully!")
        return redirect(url_for("dashboard"))

    except Exception as e:
        flash(f"Error saving image from camera: {e}")
        return redirect(url_for("upload_face"))

@app.route("/mark_attendance", methods=["GET", "POST"])
def mark_attendance():
    if "username" not in session:
        return redirect(url_for("login"))

    if session["role"] == "student":
        flash("Students are not allowed to mark attendance directly.")
        return redirect(url_for("dashboard"))
    else:
        return staff_mark_attendance()

def staff_mark_attendance():
    if "username" not in session or session["role"] != "staff":
        return redirect(url_for("login"))

    staff_subject = session.get("subject")

    if request.method == "POST":
        current_time = datetime.now()
        date_str = current_time.strftime("%Y-%m-%d")

        # Initialize webcam
        video_capture = cv2.VideoCapture(0)

        if not video_capture.isOpened():
            flash("Error: Could not access webcam.")
            return redirect(url_for("mark_attendance"))

        # Capture frame
        ret, frame = video_capture.read()
        video_capture.release()

        if not ret or frame is None:
            flash("Error: Could not capture frame from webcam.")
            return redirect(url_for("mark_attendance"))

        # Process frame
        small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
        rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)

        face_locations = face_recognition.face_locations(rgb_small_frame)
        if not face_locations:
            flash("No face detected in the frame.")
            return redirect(url_for("mark_attendance"))

        face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)
        if not face_encodings:
            flash("No face encodings found.")
            return redirect(url_for("mark_attendance"))

        recognized_students = set()

        for face_encoding in face_encodings:
            matches = face_recognition.compare_faces(known_face_encodings, face_encoding, tolerance=0.5)
            face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
            best_match_index = np.argmin(face_distances) if matches else None

            if best_match_index is not None and matches[best_match_index]:
                username = known_face_names[best_match_index]
                recognized_students.add(username)

        if not recognized_students:
            flash("No recognized students in the frame.")
            return redirect(url_for("mark_attendance"))

        # Get class name from first recognized student (assuming same class)
        first_student = Student.query.filter_by(username=list(recognized_students)[0]).first()
        if not first_student:
            flash("Student records not found.")
            return redirect(url_for("mark_attendance"))

        class_name = first_student.class_name
        attendance_file = os.path.join(app.config['ATTENDANCE_FOLDER'],
                                       f"{class_name}_{date_str}.xlsx")

        # Create or load attendance file
        if os.path.exists(attendance_file):
            wb = openpyxl.load_workbook(attendance_file)
        else:
            wb = openpyxl.Workbook()

        ws = wb.active

        # Add headers if new file
        if ws.max_row == 1:
            ws.append(["Roll Number", "Name", "Time", "Subject"])

        # Mark attendance for each recognized student
        for username in recognized_students:
            student = Student.query.filter_by(username=username).first()
            if not student:
                continue

            # Check if already marked attendance today for this subject
            already_marked = False
            for row in ws.iter_rows(values_only=True):
                if row[0] == student.roll_number and row[3] == staff_subject:
                    already_marked = True
                    break

            if not already_marked:
                ws.append([
                    student.roll_number,
                    student.full_name,
                    current_time.strftime("%H:%M:%S"),
                    staff_subject
                ])

        wb.save(attendance_file)
        flash(f"Attendance recorded for {len(recognized_students)} students in {staff_subject}")
        return redirect(url_for("dashboard"))

    return render_template("staff_mark_attendance.html")

@app.route("/view_attendance", methods=["GET", "POST"])
def view_attendance():
    if "username" not in session:
        return redirect(url_for("login"))

    if session["role"] == "student":
        return student_view_attendance()
    else:
        return staff_view_attendance()

def student_view_attendance():
    if "username" not in session or session["role"] != "student":
        return redirect(url_for("login"))

    student = Student.query.filter_by(username=session["username"]).first()
    if not student:
        flash("Student record not found.")
        return redirect(url_for("dashboard"))

    attendance_records = []
    attendance_files = [f for f in os.listdir(app.config['ATTENDANCE_FOLDER'])
                        if f.startswith(f"{student.class_name}_") and f.endswith(".xlsx")]

    print(f"Student Class: {student.class_name}")
    print(f"Considering files for {student.username}: {attendance_files}")

    for file in attendance_files:
        file_path = os.path.join(app.config['ATTENDANCE_FOLDER'], file)
        date_part = file.replace(f"{student.class_name}_", "").replace(".xlsx", "")
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            headers = [cell.value for cell in ws[1]] if ws.max_row > 1 else []
            print(f"Headers in {file}: {headers}")
            if "Roll Number" in headers and "Time" in headers and "Subject" in headers:
                roll_index = headers.index("Roll Number")
                time_index = headers.index("Time")
                subject_index = headers.index("Subject")
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[roll_index] == student.roll_number:
                        attendance_records.append({
                            "date": date_part,
                            "time": row[time_index],
                            "subject": row[subject_index]
                        })
                        print(f"Found attendance for {student.username} in {file}: {attendance_records[-1]}")
                        break  # Assuming only one entry per subject per day
            else:
                flash(f"Warning: Missing headers in {file}. Skipping.")
        except Exception as e:
            flash(f"Error reading attendance file {file}: {e}")
            print(f"Error reading {file}: {e}")

    return render_template("student_view_attendance.html", attendance=attendance_records, student=student)

def staff_view_attendance():
    if "username" not in session or session["role"] != "staff":
        return redirect(url_for("login"))

    staff_subject = session.get("subject")
    class_name = request.args.get('class_name')
    date_str = request.args.get('date')
    attendance_records = []

    if class_name and date_str and staff_subject:
        attendance_file = os.path.join(app.config['ATTENDANCE_FOLDER'], f"{class_name}_{date_str}.xlsx")
        if os.path.exists(attendance_file):
            try:
                wb = openpyxl.load_workbook(attendance_file)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    record = dict(zip(headers, row))
                    if record.get("Subject") == staff_subject:
                        attendance_records.append(record)
            except Exception as e:
                flash(f"Error reading attendance file: {e}")
        elif class_name and date_str:
            flash("Attendance file not found for the selected class and date.")
        elif not staff_subject:
            flash("No subject assigned to this staff member.")

    # Get list of unique class names for dropdown
    classes = set()
    for filename in os.listdir(app.config['ATTENDANCE_FOLDER']):
        if filename.endswith(".xlsx"):
            class_part = filename.split("_")[0]
            classes.add(class_part)

    return render_template("staff_view_attendance.html", attendance=attendance_records, classes=sorted(list(classes)))

@app.route("/logout")
def logout():
    session.pop("username", None)
    session.pop("role", None)
    session.pop("subject", None)
    flash("Logged out successfully.")
    return redirect(url_for("home"))

if __name__ == "__main__":
    with app.app_context():
        db.create_all()
        # Create a default admin user if not exists
        if not User.query.filter_by(username="admin").first():
            admin = User(username="admin", role="staff")
            admin.set_password("admin123")
            db.session.add(admin)
            db.session.commit()
    app.run(debug=True)
